def checkKeyInDictionary(searchdict, searchkey):
    #used to check if a key exists in dictionary and makes it unique if it does
    counter = 1
    newSearchKey = searchkey
    while newSearchKey in searchdict:
        #increment key as original was found
        newSearchKey = searchkey + '-' + str(counter)
        counter += 1
    return newSearchKey

def fillDictionaryFromXML (xmlroot):    
    #this block of code will fill a dictionary with the xml data for application and will return a dictionary object
    #xmlroot = xml object containing file read
    from itertools import count
    import xml.etree.ElementTree as ET
    #ns = {'app': 'http://dto.wa.ca.com/application'}
    mydict = {}     #Main Dictionary used to hold all app and job information
    myjobdict = {}  #Application or Individual job information dictionary
    hasAppRunScheduling = False
    pastJobTemplates = False

    #get application name
    if 'name' in xmlroot.attrib:
        app_name = xmlroot.attrib['name']
        myjobdict['app_name'] = app_name

    #cycle through the child elements
    for child in xmlroot:
        #get job name if we have gotten to a jobs definition
        job_name = ''
        job_qualifier = ''
        if 'name' in child.attrib:         
            job_name = child.attrib['name']
        if 'qualifier' in child.attrib:         
            job_qualifier = child.attrib['qualifier']
        if job_qualifier != '':
            job_name = job_name + "." + job_qualifier    
        #Clear job dictionary for fill we would only have a job name once we are past the applications information
        if job_name != '':
            myjobdict = {}
            myjobdict['Job_Type'] = child.tag[34:]
        #clear dictionary tag for unique naming used because there maybe elements with same naming like multipe Run schedules    
        #dicttag = ''    
        #cycle through all the elements of the child  
        for x in child.iter():
            #check for app run scheduling for job validation later
            if not pastJobTemplates and x.tag[34:] == 'schedule':
                hasAppRunScheduling =True
            #save application information if hit the job templates section of xml
            if x.tag[34:] == 'job_templates':                       
                mydict['APPLICATION'] = dict(sorted(myjobdict.items()))
                pastJobTemplates = True
                myjobdict = {}

            #if there is a value for the xml tag then save the data
            if x.text != None:
                #add child element to dictionary    
                myjobdict[checkKeyInDictionary(myjobdict, x.tag[34:])] = x.text  #add value to dictionary            
            #else:
            #    #save the parent element name for adding to dicationary as we will be drilling down into its children
            #    dicttag = x.tag[34:]
        #save job information
        if job_name != '':
                #mydict[job_name] = myjobdict #dict(sorted(myjobdict.items()))
                mydict[job_name] = dict(sorted(myjobdict.items()))

    #print('mydict=',mydict)
    return mydict, hasAppRunScheduling

def compareDictionaries (mydict, mydict2, prodfile, newfile, filename, processDependancys):
    #module will compare 2 dictionaries and print out the differences.
    #mydict = productions dictionary
    #mydict2 = new file dictionary
    #prodfile = path and file name for prod file location
    #newfile = path and file name for new file location
    #filename = just the prod file name
    #job_defaults = {'notifynodefaults', 'snmp_notifynodefaults', 'alert_notifynodefaults', 'conditional', 'criticaljob', 'subappl_wait', 'hold', 'last_notify_email','last_notify_alerts','last_notify_snmp','estimate_endtime','prop_dueout','noinherit','notrigger_ifactive','suppress_nowork_notification','reason_required','App-applopts - estimate_endtime','App-applopts - prop_dueout','App-applopts - hold','App-applopts - noinherit','App-applopts - notrigger_ifactive','App-applopts - suppress_nowork_notification','App-applopts - reason_required'} #removed 'request'
    job_defaults = {'relcount','notifynodefaults', 'snmp_notifynodefaults', 'alert_notifynodefaults', 'conditional', 'criticaljob', 'subappl_wait', 'hold', 'last_notify_email','last_notify_alerts','last_notify_snmp','estimate_endtime','prop_dueout','noinherit','notrigger_ifactive','suppress_nowork_notification','reason_required','estimate_endtime','prop_dueout','hold','noinherit','notrigger_ifactive','suppress_nowork_notification','reason_required'} #removed 'request'
    nodifferences = True
    cngdict = {}    #dictionary to hold changed jobs
    newdict = {}    #dictionary to hold new jobs added
    deldict = {}    #dictionary to hold deleted jobs
    newaddjobs = {} #dictionary to hold new job information
    keycount = 100    #used for indexing key for same job
    checkJobDependancies = False

    print('Comparing Prod File:', prodfile, ' to New File:', newfile)
    #cycle through the dictionaries comparing Orig values to New
    for key in mydict:
        #check if job is in the production file          
        if key in mydict2:                    
            if mydict[key] == mydict2[key]:
                #add the app information record for job validation
                if 'version' in mydict[key]:                    
                    newaddjobs[filename + "-" + key] = mydict2[key]
                    newaddjobs[filename + "-" + key]['JobName'] = key
                    newaddjobs[filename + "-" + key]['AppName'] = filename
                
                del mydict2[key]    #remove the key from New dicationary as they are the same            
            
            else: #cycle through all elements as there are differences
                #print('KEY=',key)
                nodifferences = False
                for elem in mydict[key]:
                    #check to see if we need to process for job dependancies
                    checkJobData = True
                    #if (('dependencies' in elem) or ('successorname' in elem) or ('relcount' in elem) or (('condition' in elem) and ('conditional' not in elem))):
                    if (('dependencies' in elem) or ('successorname' in elem) or ('relcount' in elem) or (('condition' in elem) and ('conditional' not in elem))):
                        if processDependancys:
                            checkJobData = True
                        else:
                            checkJobData = False
                    if elem in mydict2[key]:
                        #Check elem to determine if there are any similar key in the dictionary with a dash. If there is this element is a sub-set of similar elements                    
                        #If multiples of same named elements we have to create a search element if there is a - in the tag
                        usesearch = False
                        if '-' in elem:
                            searchelem = elem[:elem.find('-')]
                            resprod = dict(filter(lambda item: searchelem in item[0], mydict[key].items()))
                            usesearch = True
                        else:
                            resprod = dict(filter(lambda item: elem in item[0], mydict[key].items()))
                        #if there is more than 1 elements with the same name the we have to manipulate the values for later comparison
                        if len(resprod) > 1 and elem != 'notifynodefaults':
                            #pull from the new dictionary the block of similar elements
                            if usesearch:
                                resnew = dict(filter(lambda item: searchelem in item[0], mydict2[key].items()))
                            else:
                                resnew = dict(filter(lambda item: elem in item[0], mydict2[key].items()))
                            #Now cycle through all the prod keys compariing to the New files keys
                            for prodkey in resprod:
                                if prodkey in resnew:
                                    for newkey in resnew:
                                        #we found a match between the keys so shift the data around aligning new dict elem to the prod ones
                                        if resprod[prodkey] == resnew[newkey]:
                                            holdvalue = resnew[prodkey]
                                            resnew[prodkey] = resnew[newkey]
                                            resnew[newkey] = holdvalue
                                            mydict2[key][prodkey] = mydict2[key][newkey]                                
                                            mydict2[key][newkey] = holdvalue

                        #check for any differences for this element                        
                        if mydict[key][elem].strip() != '' and mydict2[key][elem].strip() != '':
                            #if (mydict[key][elem] != mydict2[key][elem]) and key != 'APPLICATION' and ((elem in job_defaults) and (mydict[key][elem] == 'false')) != True and (('successorname' not in elem) and ('relcount' not in elem) and ('condition' not in elem)):
                            #GOOD if (mydict[key][elem] != mydict2[key][elem]) and elem != 'app_name' and ((elem in job_defaults) and (mydict[key][elem] == 'false')) != True and (('dependencies' not in elem) and ('successorname' not in elem) and ('relcount' not in elem) and ('condition' not in elem)):                    
                            if (mydict[key][elem] != mydict2[key][elem]) and elem != 'app_name' and ((elem in job_defaults) and (mydict[key][elem] == 'false')) != True and checkJobData:                                                    
                                keycount = keycount + 1                            
                                cngdict[filename + '-' + key + ' - Change ' + str(keycount)] = 'CHANGED Parameter: ' + elem + ': Orig: ' + mydict[key][elem] + ' New: ' + mydict2[key][elem]
                                print(key, elem + ': Orig:', mydict[key][elem], ' New:', mydict2[key][elem])
                                del mydict2[key][elem] #remove the element from New dicationary as we capture difference here            
                    else:
                        #GOOD if ((elem not in job_defaults) or (elem != 'request')) and (mydict[key][elem] != 'false') and (('dependencies' not in elem) and ('successorname' not in elem) and ('relcount' not in elem) and ('condition' not in elem)):                        
                        if ((elem not in job_defaults) or (elem != 'request')) and (mydict[key][elem] != 'false') and checkJobData:                                                    
                            keycount = keycount + 1
                            cngdict[filename + '-' + key + ' - Change ' + str(keycount)] = 'CHANGED Del Param: ' + elem + ': ' + str(mydict[key][elem])
                            print(key, elem + ':', mydict[key][elem], ' DELETED-not in new record')
                
                #add the app information record for job validation
                if 'version' in mydict2[key]:                    
                    newaddjobs[filename + "-" + key] = mydict2[key]
                    newaddjobs[filename + "-" + key]['JobName'] = key
                    newaddjobs[filename + "-" + key]['AppName'] = filename
        else: #job not found in prod file
            if 'version' not in mydict[key]:
                nodifferences = False
                keycount = keycount + 1
                deldict[filename + '-' + key + ' - Change ' + str(keycount)] = 'DELETED Job'
                print(key,' job has been DELETED') #can remove from dictionary and save to another deleted dictionary            
    if nodifferences:
        print('No Differences Found in Files')

    print('')
    print('Comparing New File:', newfile, ' to Prod File:', prodfile)
    #cycle through the dictionaries comparing New values to Orig
    for key in mydict2:    
        if key in mydict:
            for elem in mydict2[key]:                
                if elem not in mydict[key]:
                    #check to see if we need to process for job dependancies
                    checkJobData = True
                    #if (('dependencies' in elem) or ('successorname' in elem) or ('relcount' in elem) or (('condition' in elem) and ('conditional' not in elem))):
                    if (('dependencies' in elem) or ('successorname' in elem) or ('relcount' in elem) or (('condition' in elem) and ('conditional' not in elem))):
                        if processDependancys:
                            checkJobData = True
                        else:
                            checkJobData = False
                    #GOOD if ((elem not in job_defaults) or (elem != 'request')) and (mydict2[key][elem] != 'false')  and (('dependencies' not in elem) and ('successorname' not in elem) and ('relcount' not in elem) and ('condition' not in elem)):                    
                    if ((elem not in job_defaults) or (elem != 'request')) and (mydict2[key][elem] != 'false')  and checkJobData:                    
                        #THIS MAYBE REMOVED CHECKING THE APPLICATION VALUE
                        #if key != 'APPLICATION' and elem not in ['JobName', 'AppName']:
                        if elem not in ['JobName', 'AppName']:
                            nodifferences = False
                            keycount = keycount + 1
                            cngdict[filename + '-' + key + ' - Change ' + str(keycount)] = 'CHANGED Add Param: ' + elem  + ': ' + mydict2[key][elem]
                            print(key, elem + ': ', mydict2[key][elem], ' ADDED-not in orig record')
        else:
            if 'version' not in mydict2[key]:
                nodifferences = False
                keycount = keycount + 1
                newdict[filename + '-' + key + ' - Change ' + str(keycount)] = 'ADDED Job'
                print(key,' job has been ADDED') #can move job information to a newjob dictionary for processing validation

                #add the new job information to new jobs dictionary for checking
                newaddjobs[filename + "-" + key] = mydict2[key]
                newaddjobs[filename + "-" + key]['JobName'] = key
                newaddjobs[filename + "-" + key]['AppName'] = filename

    if nodifferences:
        print('No Differences Found in Files')

    #print('Changed Dict =', cngdict)
    #print('New Dict =', newdict)
    #print('Del Dict =', deldict)
    #print('NewAddJobs = ',newaddjobs)
    
    return cngdict, newdict, deldict, newaddjobs

def printappchanges(totalalldict):
    lastApp = ""
    lastJob = ""
    appChangeCount = 0
    changeCount = 0
    jobDepChangeCount = 0
    delCount = 0
    addCount = 0
    depdiff = False
    appRecapDic = {}
  
    print('')
    print('Final Application Changes Listing')          
    sorted_dict = dict(sorted(totalalldict.items()))

    for key in sorted_dict:
        print(key, sorted_dict[key])
    
        keylist = key.split('-') #keylist[0]=app name, keylist[1]=job name, keylist[2]=changed
    
        if lastApp == "":
            lastApp = keylist[0]
            
        if lastApp != keylist[0]:
            #record information for the last app
            recapDic = {}
            if changeCount == 0 and delCount == 0 and addCount == 0 and jobDepChangeCount == 0:
                recapDic['Change'] = 'New App'
                recapDic['Added'] = 'New App' 
                recapDic['Deleted'] = 'New App'
                recapDic['App_Changes'] = 'New App'
                recapDic['Job_Dep_Change'] = 'New App'
            else:
                recapDic['Change'] = changeCount
                recapDic['Added'] = addCount
                recapDic['Deleted'] = delCount
                recapDic['App_Changes'] = appChangeCount
                recapDic['Job_Dep_Change'] = jobDepChangeCount
            appRecapDic[lastApp] = recapDic
            lastApp = keylist[0]
            lastJob = ""
            appChangeCount = 0
            changeCount = 0
            delCount = 0
            addCount = 0
            jobDepChangeCount = 0
            depdiff = False

        if len(keylist) > 1:            
            if "CHANGED" in sorted_dict[key]:
                if (lastJob == "") or (lastJob != keylist[1]):
                    if keylist[1] == 'APPLICATION ':
                        appChangeCount = appChangeCount + 1
                    elif ("Dependancy Differences" in sorted_dict[key]) or ("relcount: 0" in sorted_dict[key]):
                        jobDepChangeCount = jobDepChangeCount + 1 
                        depdiff = True
                    else:
                        changeCount = changeCount + 1
                elif (lastJob == keylist[1]) and depdiff and ("Dependancy Differences" not in sorted_dict[key]):
                    changeCount = changeCount + 1                        
            elif "DELETED" in sorted_dict[key]:
                delCount = delCount + 1
            elif "ADDED" in sorted_dict[key]:
                addCount = addCount + 1
            lastJob = keylist[1]

    recapDic = {}
    if changeCount == 0 and delCount == 0 and addCount == 0 and jobDepChangeCount == 0:
        recapDic['Change'] = 'New App'
        recapDic['Added'] = 'New App' 
        recapDic['Deleted'] = 'New App'
        recapDic['App_Changes'] = 'New App'
        recapDic['Job_Dep_Change'] = 'New App'
    else:
        recapDic['Change'] = changeCount
        recapDic['Added'] = addCount
        recapDic['Deleted'] = delCount
        recapDic['App_Changes'] = appChangeCount
        recapDic['Job_Dep_Change'] = jobDepChangeCount
    appRecapDic[lastApp] = recapDic

    print('')
    print('Summary Counts of Changes')
    print('Application \tChanged \tAdded \tDeleted \tApp Changes \tJob Dependancy Changes')
    for key in appRecapDic:
        print(key, '\t', appRecapDic[key]['Change'], '\t',  appRecapDic[key]['Added'], appRecapDic[key]['Deleted'], '\t', appRecapDic[key]['App_Changes'], '\t', appRecapDic[key]['Job_Dep_Change'])

    return appRecapDic

def clear():
    # define our clear function
    # import only system from os
    from os import system, name
    
    # for windows
    if name == 'nt':
        _ = system('cls')  
    # for mac and linux(here, os.name is 'posix')
    else:
        _ = system('clear')

def main():
    import xml.etree.ElementTree as ET
    import os
    import subprocess
    finalcngdict = {}       #dictionary to hold changed jobs
    finalnewdict = {}       #dictionary to hold new jobs
    finaldeldict = {}       #dictionary to hold deleted jobs
    totalalldict = {}
    newappdict = {}         #dicationary to hold new apps added
    finalnewaddjobs = {}    #dictionary to hold new job information
    comparenewfile = ''     #used to hold a trimmed down new file name
    compareprodfile = ''    #used to hold a trimmed down prod  file name
    changeJiraNumber = ''   #used to hold the Change number or Jira story for file write
    continueProcessing = True
    processDependancys = True
    prodHasAppRunScheduling = False
    newHasAppRunScheduling = False
    
    #clear the screen for fresh data
    clear()
    
    #get programming defaults
    xmlnewfilepath, xmlprodfilepath, processDependancys, printTooFile, saveFilePath, ConfigSettingsFilePath = getDefaults()
    print('')
    changeJiraNumber = input('What is the Jira story or Change Request Number?: ')
    
    #change to Prod file directory and get file list
    os.chdir(xmlprodfilepath)
    prodfiles = os.listdir(xmlprodfilepath)
    #change to New file directory and get file list
    os.chdir(xmlnewfilepath)
    newfiles = os.listdir(xmlnewfilepath)

    print('')
    print('STARTING COMPAIRISONS')
    print('*****************************************************************************')
    for prodfile in prodfiles:
        continueProcessing = True
        if prodfile[-4:] == '.xml': #remove the end section
            compareprodfile = prodfile[0:-4]
        else:
            compareprodfile = prodfile  #6/28/22-Changed was compareprodfile

        for newfile in newfiles:            
            if (newfile[0:5].upper() == 'TEST_') or (newfile[0:5].upper() == 'VDEV_'):     #remove the prefix
                comparenewfile = newfile[5:]                
            elif newfile[0:4].upper() == 'E2E_':
                comparenewfile = newfile[4:]                
            else:
                comparenewfile = newfile
                
            if comparenewfile[-4:] == '.xml': #remove the end section
                comparenewfile = comparenewfile[0:-4]

            #print('prodfile=',compareprodfile)
            #print('newfile=',comparenewfile)
            #print('leng prod=',len(compareprodfile))
            #print('leng new=',len(comparenewfile))
            #print('length=',comparenewfile[len(compareprodfile):len(compareprodfile) + 1])
            if continueProcessing:
                #if compareprodfile == comparenewfile[0:len(compareprodfile)]:                                
                if ((compareprodfile == comparenewfile[0:len(compareprodfile)]) and (len(compareprodfile) == len(comparenewfile))) or ((compareprodfile == comparenewfile[0:len(compareprodfile)]) and (comparenewfile[len(compareprodfile):len(compareprodfile) + 1] == '_')):                
                    #fill prod dictionary            
                    mytree = ET.parse(os.path.join(xmlprodfilepath, prodfile))
                    root = mytree.getroot()
                    proddict, prodHasAppRunScheduling = fillDictionaryFromXML(root)
                    
                    #fill new dictionary            
                    mytree = ET.parse(os.path.join(xmlnewfilepath, newfile))
                    root = mytree.getroot()    
                    newdict, newHasAppRunScheduling = fillDictionaryFromXML(root)                          
                    #print('proddict=',proddict)
                    #print('newdict=',newdict)
                    #compare the dictionaries  
                    cngdict, newdict, deldict, newaddjobs = compareDictionaries(proddict, newdict, xmlprodfilepath + '\\' + prodfile, xmlnewfilepath + '\\' + newfile, prodfile, processDependancys)
                    
                    #accumulate all changes into 1 file
                    finalcngdict.update(cngdict)
                    finalnewdict.update(newdict)                    
                    finaldeldict.update(deldict)
                    finalnewaddjobs.update(newaddjobs)

                    #remove files processed
                    newfiles.remove(newfile)                    
                    continueProcessing = False

        if continueProcessing:        
            print(newfile, 'File not found in PROD folder')
        print('*****************************************************************************')

    for file in newfiles:
        print('NEW Application, not found in Prod file folder:', file)        
        #Accumulate the jobs to newjobs dictionary                    
        mytree = ET.parse(os.path.join(xmlnewfilepath, file))
        root = mytree.getroot()    
        newdict, newHasAppRunScheduling = fillDictionaryFromXML(root)
                
        newaddjobs = {}
        for job in newdict:
            #add the new job information to new jobs dictionary for checking            
            newaddjobs[file + "-" + job] = newdict[job]                      
            newaddjobs[file + "-" + job]['AppName'] = file
            newaddjobs[file + "-" + job]['JobName'] = job
            if "Job_Type" not in newdict[job]:
                newaddjobs[file + "-" + job]['Job_Type'] = 'New Application'

        #add description to changes information for the new application
        newappdict[file] = 'NEW Application'
        #newappdict['Job_Type'] = 'NEW Application'
        finalnewdict.update(newappdict)

        finalnewaddjobs.update(newaddjobs)
    
    print('*****************************************************************************')
    
    #accumlate all changes into 1 dictionary
    totalalldict.update(finalcngdict)
    totalalldict.update(finalnewdict)
    totalalldict.update(finaldeldict)
    #print final result
    recapAllChangesDic = printappchanges(totalalldict)

    #run job checking routine    
    finalnewaddjobs = checknewjobsdata(finalnewaddjobs,ConfigSettingsFilePath,newHasAppRunScheduling)

    #print report
    if printTooFile == 'Excel':
        #print to excel final changes for differences
        printexcelfile(totalalldict, 1, changeJiraNumber, saveFilePath)
        printexcelfile(finalnewaddjobs, 2, changeJiraNumber, saveFilePath)
        printexcelfile(recapAllChangesDic, 3, changeJiraNumber, saveFilePath)
    else:
        #print to text file
        printtextfile(totalalldict, 1, changeJiraNumber, saveFilePath)    
        printtextfile(finalnewaddjobs, 2, changeJiraNumber, saveFilePath)
        printtextfile(recapAllChangesDic, 3, changeJiraNumber, saveFilePath)
    
    #Check if user wants to open excel results file
    print('')
    print('')
    if printTooFile == 'Excel':
        workbookfile = os.path.join(saveFilePath + '\\Comparison_Results-' + changeJiraNumber + '.xlsx')
        openFile = input('Open Excel File ' + workbookfile + ' Yes(Y):').upper()
    else:
        workbookfile = os.path.join(saveFilePath + '\\Comparison_Results-' + changeJiraNumber + '.txt')
        openFile = input('Open Text File ' + workbookfile + ' Yes(Y):').upper()    
    if openFile == 'Y':        
        if printTooFile == 'Excel':            
            os.startfile(workbookfile)  #Using this one to launch the fiel, other tries below keep the app frozen until you close the excel file.
            #os.system("start EXCEL.EXE " + workbookfile)
            #subprocess.call(workbookfile,shell = True)            
            #p = subprocess.Popen(workbookfile, shell=True)
            #p.terminate()
        else:
            os.startfile(filepath=workbookfile)
    
    #print('finaladdjobs=', finalnewaddjobs)
    #print('finalcngdict=', finalcngdict)
    #print('finaldeldict=', finaldeldict)
    os.close

def printexcelfile(totalalldict, changetype, changeJiraNumber, saveFilePath):
    #totalalldict = dictionary of changes coming in to print
    #changetype(1=All Differences,2=New Jobs,3=Summary counts)
    #changeJiraNumber = Used in construction of file saving    
    #saveFilePath = path to where results file will be saved

    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter  #utility to get column letters.
    from openpyxl.styles import Font, PatternFill    
    import re

    #used to color row for task/link jobs to green
    colornextrowgreen = False
    colornextrowblue = False      
    colornextrowpurple = False       
    linktaskjob = False
    externaljob = False
    applicationrow = False
    fillcolorgreen = PatternFill(patternType='solid', fgColor='D4FFD1')
    fillcolordkgreen = PatternFill(patternType='solid', fgColor='006400')
    fillcolorred = PatternFill(patternType='solid', fgColor='FFDDD1')
    fillcolorltblue = PatternFill(patternType='solid', fgColor='D1EFFF')
    fillcolorltpurple = PatternFill(patternType='solid', fgColor='E6BBF6')
    fillcoloryellow = PatternFill(patternType='solid', fgColor='FFF700')
    
    workbookfile = saveFilePath + '\\Comparison_Results-' + changeJiraNumber + '.xlsx'

    if changetype == 1:
        #create new empty work book
        wb = Workbook()
    else:
        #load existing work to add new jobs information
        wb = load_workbook(workbookfile)

    #set to first sheet to begin working on
    ws = wb.active

    ws.title = "App Differences"    #name the sheet
    headings = []    
    if changetype == 1:        
        headings.append('Final Application Changes Listing')
        ws.append(headings)              
        sorted_dict = dict(sorted(totalalldict.items()))

        for key in sorted_dict:
            headings = []            
            #headings = [key] + [sorted_dict[key]]
            if ' - Change' in key:
                headings = [key[0:key.index(' - Change')]] + [sorted_dict[key]]            
            else:
                headings = [key] + [sorted_dict[key]]            
            ws.append(headings)
    elif changetype == 2:
        ws.append([])        
        ws.append([])
        ws.append(['NEW JOBS TO VALIDATE'])                
        ws.append(['    APP INFORMATION'])                
        ws.append(['     LINK/TASK JOBS'])                
        ws.append(['     EXTER MON JOBS'])                
        sorted_dict = dict(sorted(totalalldict.items()))
    
        for key in sorted_dict: #cycle through all rows in the added jobs dictionary
            #if 'INVALID_ENTRIES' in totalalldict[key]:
            headings = ['AppName', 'JobName', 'INVALID_ENTRIES'] + list(totalalldict[key].keys())                                          #create column headings using the data. all keys same            
            ws.append(headings)                                                                                         #add the heading to sheet            
            grades = [totalalldict[key]['AppName'], totalalldict[key]['JobName'], totalalldict[key]['INVALID_ENTRIES']] + list(totalalldict[key].values())    #get all the grades as a list for each person
            ws.append(grades)                                                                                           #append to the excel sheet each row of data
            ws.append([])                                                                                               #add a blank line
            #else:
            #    headings = ['AppName', 'JobName'] + list(totalalldict[key].keys())                                          #create column headings using the data. all keys same            
            #    ws.append(headings)                                                                                         #add the heading to sheet            
            #    grades = [totalalldict[key]['AppName'], totalalldict[key]['JobName']] + list(totalalldict[key].values())    #get all the grades as a list for each person
            #    ws.append(grades)                                                                                           #append to the excel sheet each row of data
            #    ws.append([])                                                                                               #add a blank line                
    else:
        ws.append([])        
        ws.append([])
        ws.append(['Summary Counts of Changes'])                                
        headings = ['Application', 'Changed', 'Added', 'Deleted', 'App Changes', 'Job Dependancy Changes']          #create column headings
        ws.append(headings)                                                                                         #add the heading to sheet            
        for key in totalalldict: #cycle through all rows in the summary dictionary            
            data = [key, totalalldict[key]['Change'], totalalldict[key]['Added'], totalalldict[key]['Deleted'], totalalldict[key]['App_Changes'], totalalldict[key]['Job_Dep_Change']]          
            ws.append(data)                                                                                           #append to the excel sheet each row of data

    #save workbook
    #wb.save(workbookfile) - original save
    import traceback
    trySaveFile = True
    while trySaveFile:
        try:
            #print('saved')
            wb.save(workbookfile)                                                   
            trySaveFile = False            
        except IOError as err:
            print('')
            print('ERROR INFORMATION:')
            traceback.print_exc()
            print('')
            hold = input('WRITE PERMISSION TO FILE, CLOSE FILE AND TRY AGAIN.  Enter any key after closing file.')
        except Exception:
            print('')
            print('ERROR INFORMATION:')
            traceback.print_exc()        

    #always close the workbook
    wb.close

    #lets add some color and bolding
    wb = load_workbook(workbookfile)
    ws = wb.active
    for row in ws.iter_rows():         
        makecellbold = False      
        
        if colornextrowgreen:
            linktaskjob = True
        else:
            linktaskjob = False

        if colornextrowblue:
            externaljob = True
        else:
            externaljob = False
        
        if colornextrowpurple:
            applicationrow = True
        else:
            applicationrow = False

        colornextrowgreen = False
        colornextrowblue = False       
        colornextrowpurple = False

        for cell in row:

            if cell.column == 2 and ws[cell.coordinate].value in ['NEW Application', 'ADDED Job']:
                ws[str(cell.coordinate)].font = Font(bold='true', color='006400')
            if cell.column == 2 and ws[cell.coordinate].value in ['CHANGED Add Param: donotrun: true', 'CHANGED Add Param: request: true', 'CHANGED Parameter: request: Orig: false New: true']:
                ws[str(cell.coordinate)].fill = fillcolorred
                ws[str(cell.coordinate)].font = Font(bold='true', color='00FF0000')                
            elif cell.column == 2 and ws[cell.coordinate].value == 'DELETED Job':
                ws[str(cell.coordinate)].font = Font(bold='true', color='0B22F6')
                
            #if its column 1 and has 1 of those values then make every cell in this row bold        
            if cell.column == 1 and (ws[cell.coordinate].value == 'AppName' or ws[cell.coordinate].value == 'JobName'):            
                makecellbold = True
            if ws[cell.coordinate].value in ['Final Application Changes Listing','Summary Counts of Changes','Application','Changed','Added','Deleted','App Changes','Job Dependancy Changes']:
                ws[str(cell.coordinate)].font = Font(bold='true')
            if ws[cell.coordinate].value == 'NEW JOBS TO VALIDATE':
                ws[str(cell.coordinate)].fill = fillcoloryellow
                ws[str(cell.coordinate)].font = Font(bold='true')
            if ws[cell.coordinate].value == '    APP INFORMATION':
                ws[str(cell.coordinate)].fill = fillcolorltpurple
                ws[str(cell.coordinate)].font = Font(bold='true')
            if ws[cell.coordinate].value == '     LINK/TASK JOBS':                
                ws[str(cell.coordinate)].fill = fillcolorgreen
                ws[str(cell.coordinate)].font = Font(bold='true')
            if ws[cell.coordinate].value == '     EXTER MON JOBS':
                ws[str(cell.coordinate)].fill = fillcolorltblue  
                ws[str(cell.coordinate)].font = Font(bold='true')              
            #if jobvalidation and cell.column == 3 and ws[cell.coordinate].value not in ['NONE', 'LINK/TASK JOBS']:
            if cell.column == 3 and ws[cell.coordinate].value not in ['NONE','Added','New App'] and cell.data_type != 'n':                
                ws[str(cell.coordinate)].font = Font(bold='true', color='00FF0000')
            if (cell.column == 3) and (ws[cell.coordinate].value == 'Job_Type') or ((cell.column == 4) and (ws[cell.coordinate].value == 'Job_Type')):
                checkposforjobNUM = re.findall("\d+", cell.coordinate)[0]                
                if (ws['D' + str(int(checkposforjobNUM) + 1)].value in ['link','task']): # or (ws['C' + str(int(checkposforjobNUM) + 1)].value in ['link','task']):
                    colornextrowgreen = True
                if (ws['D' + str(int(checkposforjobNUM) + 1)].value in ['ext_job','extmon_job']): 
                    colornextrowblue = True
            #if (cell.column == 5) and (ws[cell.coordinate].value == 'version'):
            if (ws[cell.coordinate].value == 'version'):
                colornextrowpurple = True
            #continue making all other cells in the header row bold
            if makecellbold:
                ws[str(cell.coordinate)].font = Font(bold='true')
            if linktaskjob:                
                ws[str(cell.coordinate)].fill = fillcolorgreen
            if externaljob:
                ws[str(cell.coordinate)].fill = fillcolorltblue
            if applicationrow:
                ws[str(cell.coordinate)].fill = fillcolorltpurple

    #save final workbook
    wb.save(workbookfile)

    #save final workbook
    #trySaveFile = True
    #while trySaveFile:
    #    try:
    #        wb.save(workbookfile)            
    #        print('Successfully saved')
    #        trySaveFile = False
    #    except IOError as err:
    #        hold = input('Write permissions to file, close file and try again.  Enter any key after closing file.')
    #        #continue  

    #always close the workbook
    wb.close

def printtextfile(totalalldict, changetype, changeJiraNumber, saveFilePath):
    #totalalldict = dictionary of changes coming in to print
    #changetype(1=All Differences,2=New Jobs,3=Summary counts)
    #changeJiraNumber = Used in construction of file saving    
    #saveFilePath = path to where results file will be saved
    
    import csv
    workbookfile = saveFilePath + '\\Comparison_Results-' + changeJiraNumber + '.txt'
    #workbookfile = 'C:\\CAWACompare\\Results\\Comparison_Results-' + changeJiraNumber + '.txt'    

    sorted_dict = dict(sorted(totalalldict.items()))    
    if changetype == 1:
        with open(workbookfile, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            xml_data = []
            xml_data.append('Final Application Changes Listing')
            writer.writerow(xml_data)            
            for key in sorted_dict:
                xml_data = []
                xml_data.append(key)
                xml_data.append(sorted_dict[key])
                writer.writerow(xml_data)
            writer.writerow('')
            writer.writerow('')
            xml_data = []
            xml_data.append('NEW JOBS TO VALIDATE')
            writer.writerow(xml_data)
    elif changetype == 2:        
        with open(workbookfile, 'a', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            for key in sorted_dict:
                xml_key_data = []
                xml_key_value = []
                firstelement = True            
                for element in sorted_dict[key]:
                    if firstelement:
                        xml_key_data.append('AppName')
                        xml_key_value.append(sorted_dict[key]['AppName']) 
                        xml_key_data.append('JobName')
                        xml_key_value.append(sorted_dict[key]['JobName'])
                        xml_key_data.append('INVALID_ENTRIES')
                        xml_key_value.append(sorted_dict[key]['INVALID_ENTRIES'])
                        firstelement = False

                    xml_key_data.append(element)
                    xml_key_value.append(sorted_dict[key][element])

                writer.writerow(xml_key_data)
                writer.writerow(xml_key_value)
                writer.writerow('')
    else: #write summary section to file
        with open(workbookfile, 'a', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            xml_key_data = ['Summary Counts of Changes']
            xml_key_value = ['Application', 'Changed', 'Added', 'Deleted', 'App Changes', 'Job Dependancy Changes']
            writer.writerow(xml_key_data)
            writer.writerow(xml_key_value)            
            for key in sorted_dict:
                xml_key_value = []
                firstelement = True
                xml_key_value.append(key)
                for element in sorted_dict[key]:
                    xml_key_value.append(sorted_dict[key][element])                    
                writer.writerow(xml_key_value)
                
    #always close the file
    csvfile.close()

def getfilepath(filePath, fileSource):    
    newFilePath = filePath
    if fileSource == 'New':
        fileSourcePath = 'New File Path'        
    elif fileSource == 'Prod':
        fileSourcePath = 'Prod File Path'
    elif fileSource == 'Save':
        fileSourcePath = 'Save Results File Path'
    elif fileSource == 'Config':
        fileSourcePath = 'Save Configuration Files Path'
    
    filePathAnswer = input('Current ' + fileSourcePath + ' is set too: ' + filePath + ' Keep this path? (Y) Change this path? (C):')
    while (filePathAnswer.upper() != 'Y'):
        filePathAnswer = input('What is your ' + fileSourcePath + ' location?:')
        newFilePath = filePathAnswer
        filePathAnswer = input('Current ' + fileSourcePath + ' is set too: ' + newFilePath + ' Keep this path? (Y) Change this path? (C):')
        
    #return the file path
    return newFilePath

def getDefaults():
    # importing the module
    import json
    import os
    changeOption = "displaydefaults"
    data = {}
    validData = False

    #set file path using location of python script as starting point
    script_path = os.path.abspath(__file__) # i.e. /path/to/dir/CAWACompareUtility.py
    script_dir = os.path.split(script_path)[0] #i.e. /path/to/dir/
    # using getlogin() returning username
    user = os.getlogin()
    rel_path = "bin\\CAWACompareSettings-" + user + ".json"
    file_path = os.path.join(script_dir, rel_path) #Final result = <CAWACompareUtility.py-Location>/bin/CAWACompareSettings-<userid>.json

    # get the default information
    try:
        with open(file_path) as json_file:
            data = json.load(json_file)
            ConfigSettingsFilePath = data['ConfigSettingsFilePath']
            newFilePath = data['newFilesFolder']
            prodFilePath = data['prodFilesFolder']
            dependancycheck = data['CheckDependanys']
            printTooFile = data['FilePrintType']
            saveFilePath = data['FileSavePath']
    except FileNotFoundError:        
        #fill in defaults
        ConfigSettingsFilePath = script_dir + '\\bin'
        newFilePath = 'C:\DummyNewFilesPath'
        prodFilePath = 'C:\DummyProdFilesPath'
        dependancycheck = False
        printTooFile = 'Excel'
        saveFilePath = 'C:\DummyResultFilesPath'
        data['ConfigSettingsFilePath'] = ConfigSettingsFilePath
        data['newFilesFolder'] = newFilePath
        data['prodFilesFolder'] = prodFilePath
        data['CheckDependanys'] = dependancycheck
        data['FilePrintType'] = printTooFile
        data['FileSavePath'] = saveFilePath
    
    while (not validData) or (changeOption != ""):        
        #clear the screen for fresh data
        clear()
        #print('file_path=',file_path)
        if (newFilePath == 'C:\DummyNewFilesPath') or (prodFilePath == 'C:\DummyProdFilesPath') or (saveFilePath == 'C:\DummyResultFilesPath'):
            print('NOTE: YOU MUST CHANGE THE FILE PATHS TO YOUR INSTALLATION CHOICES')        
        #Display all Options Using
        print('Default Options Using')        
        print('1. New      Files Folder: ', newFilePath)
        print('2. Prod     Files Folder: ', prodFilePath)
        print('3. Results  Files Folder: ', saveFilePath)
        print('4. Config   Files Folder: ', ConfigSettingsFilePath)
        print('5. Check For Dependancys: ', dependancycheck)
        print('6. Save as (Excel/Text) : ', printTooFile)
        print('')
        changeOption = input('These are your default settings.  If you like to change any settings enter selection Number or Enter to Use defaults:')

        if changeOption == '1':
            # Check for new file path
            newFilePath = getfilepath(newFilePath, 'New')
        if changeOption == '2':
            # Check for prod file path
            prodFilePath = getfilepath(prodFilePath, 'Prod')
        if changeOption == '3':
            saveFilePath = getfilepath(saveFilePath, 'Save')
        if changeOption == '4':
            ConfigSettingsFilePath = getfilepath(ConfigSettingsFilePath, 'Config')
        if changeOption == '5':
            # Check for job naming standards
            processDependancys = input('Check Jobs Dependancys is set too (' + str(data['CheckDependanys']) + ')  Keep this setting? (Y) Change this setting? (C):')
            if processDependancys in ['Y', 'y']:
                dependancycheck = data['CheckDependanys']
            else:
                dependancycheck = not data['CheckDependanys']
        if changeOption == '6':
            # Check for file print type
            if printTooFile == 'Excel':
                processDependancys = input('Results printed to (' + data['FilePrintType'] + ') File.  Keep this setting? (Y) or Print to Text File? (T):')
                if processDependancys in ['Y', 'y']:
                    printTooFile = data['FilePrintType']
                else:
                    printTooFile = 'Text'
            else:
                processDependancys = input('Results printed to (' + data['FilePrintType'] + ') File.  Keep this setting? (Y) or Print to Excel File? (E):')
                if processDependancys in ['Y', 'y']:
                    printTooFile = data['FilePrintType']
                else:
                    printTooFile = 'Excel'
        if (ConfigSettingsFilePath == 'C:\DummySettingsPath') or (newFilePath == 'C:\DummyNewFilesPath') or (prodFilePath == 'C:\DummyProdFilesPath') or (saveFilePath == 'C:\DummyResultFilesPath'):
            validData = False            
        else:
            validData = True
        
    # if paths changed update back to json file
    if (newFilePath != data['newFilesFolder']) or (prodFilePath != data['prodFilesFolder']) or (saveFilePath != data['FileSavePath']) or (ConfigSettingsFilePath != data['ConfigSettingsFilePath']) or (dependancycheck != data['CheckDependanys']) or (printTooFile != data['FilePrintType']):
        data['newFilesFolder'] = newFilePath
        data['prodFilesFolder'] = prodFilePath
        data['FileSavePath'] = saveFilePath
        data['ConfigSettingsFilePath'] = ConfigSettingsFilePath
        data['CheckDependanys'] = dependancycheck
        data['FilePrintType'] = printTooFile
        
        with open(ConfigSettingsFilePath + "\\CAWACompareSettings-" + user + ".json", "w") as outfile:
            json.dump(data, outfile)
            #MUDDU-should we close this object before exiting the module? outfile(close)    
    return newFilePath, prodFilePath, dependancycheck, printTooFile, saveFilePath, ConfigSettingsFilePath

def checknewjobsdata(testdic, ConfigSettingsFilePath,newHasAppRunScheduling):
    import os
    jobname = ''
    knownJobTypes = ['LINK','TASK','EXTMON_JOB','EXT_JOB','SFTP_JOB','UNIX_JOB','FILEMON_JOB','ORACLE_APPS_REQUEST_SET','ORACLE_APPS_SINGLE_REQUEST','PEOPLESOFT_JOB','NT_JOB','MSSQLSERVER_JOB']
    
    #set file path using location of python script as starting point
    script_path = os.path.abspath(__file__) # i.e. /path/to/dir/CAWACompareUtility.py
    script_dir = os.path.split(script_path)[0] #i.e. /path/to/dir/
    # using getlogin() returning username
    user = os.getlogin()
    rel_path = "SubSystemsList.txt"
    file_path = os.path.join(script_dir, rel_path) #Final result = <CAWACompareUtility.py-Location>/bin/CAWACompareSettings-<userid>.json
    #print('subsystem=',file_path)
    #get sub-system listing
    List = open(file_path).read().splitlines()
    #List = open(ConfigSettingsFilePath + "\\SubSystemsList.txt").read().splitlines()
    #List = open(os.path.join(ConfigSettingsFilePath, 'SubSystemsList.txt')).read().splitlines() 07/13/22 ADDED RESEARCH

    #MUDDU-should we close this object before exiting the module? List(close)
    #print('CHECKTESTDIC=',testdic)

    for key in testdic:
        #print(testdic)            
        validJobName = True
        validOmniJobName = False
        errormsg = ''        
        if ('version' in testdic[key]) or (testdic[key]['Job_Type'].upper() == 'LINK') or (testdic[key]['Job_Type'].upper() == 'TASK'):                            
            testdic[key]['INVALID_ENTRIES'] = 'NONE'
        elif (testdic[key]['Job_Type'].upper() == 'EXT_JOB') or (testdic[key]['Job_Type'].upper() == 'EXTMON_JOB'):
            #place holder, add check for the scheduled limits
            #<app:scheduled>00:00 TODAY</app:scheduled> From value
            #<app:scheduledlimit>00:00 TOMORROW</app:scheduledlimit> Too Value
            if ('scheduled' not in testdic[key]) or ('scheduledlimit' not in testdic[key]):
                errormsg = errormsg + ' Invalid Scheduled Limits'
        else:
            #get the job name without qualifier
            if "." in testdic[key]['JobName']:
                jobname = testdic[key]['JobName'][:testdic[key]['JobName'].find('.')]
            else:    
                jobname = testdic[key]['JobName']
            # Omni name job validation
            if jobname[0:4].upper() in ['PAM_','PPC_']:
                validOmniJobName = True
            elif jobname[0:5].upper() in ['PAXE_','PICM_','PMDS_','PEMR_','PESP_','PINV_','PMCM_','POMS_','PPAL_','PPUB_','PSDC_','PWMS_']:
                validOmniJobName = True
            elif jobname[0:6].upper() in ['BOPIS_','PCUST_','PITEM_','PXREF_']:
                validOmniJobName = True
            elif jobname[0:7].upper() in ['PACMAN_','PCLEAR_','PHBASE_','PNGPOS_','PPURGE_','PTOOLS_','PTRANS_']:
                validOmniJobName = True
            elif jobname[0:8].upper() in ['PREPORT_','PSELECT_','PROCESS_']:
                validOmniJobName = True
            elif jobname[0:9].upper() in ['PAUTOPOP_','PCATALOG_','PHYSICAL_','PPRICECA_','PPRICEJP_','PPRICEUK_','PPRICING_','PPRODUCT_','PPUBLISH_']:
                validOmniJobName = True
            elif jobname[0:11].upper() in ['POSRECLAIM_']:
                validOmniJobName = True
            
            if not validOmniJobName:
                # check for prod
                if jobname[0:1].upper() != 'P':
                    errormsg = 'First letter of Job name not P'

                # check for subsystem name
                if jobname[1:4] not in List:
                    errormsg = errormsg + ' Invalid SubSystem Name'

                # check for frequency: Frequency (D-Daily, H-Hourly, W-Weekly, M-monthly, R-Request etc)                
                if jobname[len(jobname)-2] in ['M','Q','Y']:
                    errorfound = True
                    for jobelem in testdic[key]:
                        if jobname[len(jobname)-2] == 'M' and 'MONTH' in jobelem.upper():
                            errorfound = False
                            #print('jobelem=',jobelem)                           
                        elif jobname[len(jobname)-2] == 'Q' and (('QUARTER' in jobelem.upper()) or ('QTR' in jobelem.upper())):
                            errorfound = False
                            #print('jobelem=',jobelem)                           
                        elif jobname[len(jobname)-2] == 'Y' and 'YEAR' in jobelem.upper():
                            errorfound = False
                            #print('jobelem=',jobelem)                           
                    if errorfound:
                        errormsg = errormsg + ' Invalid Frequency'
                elif jobname[len(jobname)-2] not in ['D','H','W','R','U']: 
                    errormsg = errormsg + ' Invalid Frequency'

                if testdic[key]['Job_Type'].upper() in knownJobTypes:
                    #check for Last Character: Platform (U-Unix, L-Linux, O-Oracle, R–Request, P-Peoplesoft, N/W-Windows, S-MSSQL Job, ? question about these UP-SFTP Upload, DW-SFTP Download etc.)
                    if (jobname[-2:].upper() in ['UP', 'DW']) and (testdic[key]['Job_Type'].upper() != 'SFTP_JOB'):
                        validJobName = False
                    #elif (jobname[-1].upper() in ['U', 'L', 'R']) and (jobname[-2].upper() not in ['UP', 'DW']) and (testdic[key]['Job_Type'].upper() != 'UNIX_JOB'):    
                    elif (jobname[-1].upper() in ['U', 'L', 'R']) and (jobname[-2].upper() not in ['UP', 'DW']) and (testdic[key]['Job_Type'].upper() not in ['UNIX_JOB','FILEMON_JOB']):                    
                        validJobName = False
                    elif (jobname[-1].upper() == 'O') and ((testdic[key]['Job_Type'].upper() != 'ORACLE_APPS_REQUEST_SET') and (testdic[key]['Job_Type'].upper() != 'ORACLE_APPS_SINGLE_REQUEST')):
                        validJobName = False
                    elif (jobname[-1].upper() == 'P') and (jobname[-2:].upper() not in ['UP', 'DW']) and (testdic[key]['Job_Type'].upper() != 'PEOPLESOFT_JOB'):
                        validJobName = False
                    elif ((jobname[-1].upper() in ['N', 'W']) and (testdic[key]['Job_Type'].upper() != 'NT_JOB')) and (jobname[-2:].upper() not in ['UP', 'DW']):
                        validJobName = False
                    elif (jobname[-1].upper() == 'S') and (testdic[key]['Job_Type'].upper() != 'MSSQLSERVER_JOB'):                
                        validJobName = False
                    elif (jobname[-1].upper() not in ['U', 'L', 'R','O','P','N','W','S']):
                        validJobName = False
                    if not validJobName:    
                        errormsg = errormsg + ' Invalid Platform'
            
        if 'donotrun' in testdic[key]:
            errormsg = errormsg + ' Job has donotrun setting'
            #print('value=',testdic[key]['donotrun'])
       
        if 'request' in testdic[key]:
            if testdic[key]['request'] == 'true':
                errormsg = errormsg + ' Job has Request setting'

        #if (currentAppName == '') and ('APPLICATION' in key):
        #    keylist = key.split('-') #keylist[0]=app name, keylist[1]=job name, keylist[2]=changed
        #    if currentAppName != keylist[0]:
        #        currentAppName = keylist[0]
        #        appHasRunScheduling = False        
        #print('KEY=',key)
        #check for run scheduling
        if ('APPLICATION' in key):
            keylist = key.split('-') #keylist[0]=app name, keylist[1]=job name, keylist[2]=changed
            if keylist[1] == 'APPLICATION':
                appHasRunScheduling = False
        jobHasRunScheduling = False        
        for jobelem in testdic[key]:            
            if jobelem == 'schedule':
                if 'APPLICATION' in key:
                    if keylist[1] == 'APPLICATION':
                        appHasRunScheduling = True
                    else:
                        jobHasRunScheduling = True        
                else:
                    jobHasRunScheduling = True        
        if (jobHasRunScheduling == False) and (appHasRunScheduling == False):        
            if 'APPLICATION' in key:
                errormsg = errormsg + ' App has NO Run Scheduling'
            else:

                errormsg = errormsg + ' Job has NO Run Scheduling'

        if errormsg != '':
            testdic[key]['INVALID_ENTRIES'] = errormsg
        elif errormsg == '':
            testdic[key]['INVALID_ENTRIES'] = 'NONE'
        elif not validOmniJobName:
            testdic[key]['INVALID_ENTRIES'] = 'Invalid Omni Job Name'            

    return testdic

main()


def fillDictionaryFromXML (xmlroot, dependancycheck):
    #this block of code will fill a dictionary with the xml data for application and will return a dictionary object
    #xmlroot = xml object containing file read
    #dependancycheck = option to check teh jobs dependancy settings    
    mydict = {}         #setup empty dictionary to story entire list of jobs
    myjobdict = {}      #job dictionary for fill
    job_name = ''       #used to hold job name
    job_qualifier = ''  #used to hold the jobs qualifier if any
    job_type = {'link','unix_job','extmon_job','ext_job','filemon_job','nt_job','oracle_apps_single_request','oracle_apps_request_set','peoplesoft_job','mssqlserver_job','task','om_textlog_job','servlet_job','zos_job','ftp_job','sftp_job'}

    #get application name
    appname = xmlroot.attrib['name']
    if appname[0:5].upper() == 'TEST_':     #remove the prefix
        appname = appname[5:]
    needappdata = True
    varcount = 0        #used for setting key index unique for common parameters
    for x in xmlroot: #cycle through base elements    
        #note using [34:] to get past "{http://dto.wa.ca.com/application}" section in tag        
        if x.tag[34:] in job_type:
            #build the key for dictionary using job name.job qualifer
            if 'qualifier' in x.attrib:         
                job_qualifier = x.attrib['qualifier']
            else:
                job_qualifier = ''
            job_name = x.attrib['name'] + '.' + job_qualifier            

            myjobdict = {}                      #Clear job dictionary for fill
            myjobdict['Job_Type'] = x.tag[34:]  #add job type
            for y in x:                         #cycle through child elements of job
                #if tag is resourcedependencies we need to collect the resource information
                if y.tag[34:] == 'resourcedependencies': 
                    for z in y:                          
                        if z.tag[34:] == 'resources':    #when child element is resources gather information
                            for w in z:                            
                                myjobdict[w[0].text + '-' + w[1].text] = 'Resource Differences'                                   
                #if tag is variable_dependencies we need to collect the information
                elif y.tag[34:] == 'variable_dependencies': 
                    for z in y:                          
                        varcount = 0
                        if z.tag[34:] == 'variable_expression':    #when child element is resources gather information                            
                            varcount = varcount + 1
                            for w in z:                                
                                myjobdict['variable_expression' + str(varcount) + ' - ' + w.tag[34:]] = w.text
                        else:
                            myjobdict['variable_expression - ' + z.tag[34:]] = z.text
                #if tag is states we need to collect information for them
                elif y.tag[34:] == 'states': 
                    for z in y:
                        if z.tag[34:] == 'exec':                        
                            for w in z:
                                myjobdict['Job exec - ' + w.tag[34:]] = w.text
                        elif z.tag[34:] == 'ready':                        
                            for w in z:
                                if w.tag[34:] == 'on_enter':                                    
                                    myjobdict['Job Script-At Job Run Time:'] = w[0].text
                                else:
                                    myjobdict['Job ready - ' + w.tag[34:]] = w.text                
                #if tag is schedules we need to collect information for them
                elif y.tag[34:] in ['schedules']:                     
                    for z in y:
                        for w in z:                                
                            myjobdict[z.tag[34:] + '-' + w.text] = '| Job ' + z.tag[34:] + ' Differences'                        
                #if tag is time_actions we need to collect information for them
                elif y.tag[34:] in ['time_actions']:                     
                    for z in y:
                        for w in z:                                
                            #myjobdict[z.tag[34:] + '-' + w.text] = '| Job ' + z.tag[34:] + w.text + ' Differences'                        
                            #myjobdict[z.tag[34:]] = '| Job ' + z.tag[34:] + ' ' + w.text + ' Differences'                                                
                            myjobdict[z.tag[34:]] = w.text
                #if tag is retry, time_actions or fileexpand we need to collect information for them
                elif y.tag[34:] in ['retry','fileexpand','psopr','email','outdestination','job_object_params','on_load','max_min']: #removed 'dependencies', 06/24/22-Removed ,'time_actions'
                    for z in y:                        
                        if z.tag[34:] in ['script_reference','script_definition']:                            
                            myjobdict['Job Script-At Event Trigger Time:'] = z.text
                        else:                            
                            myjobdict[z.tag[34:]] = z.text
                #Add coding for emailing capture
                elif y.tag[34:] in ['notifylist','alert_notifylist','snmp_notifylist']:                                        
                    for z in y:
                        emailkey = ''                                                                        
                        for w in z:
                            if w.tag[34:] in ['monitor_states', 'mailaddresslist']:
                                for ww in w:                                    
                                    emailkey = emailkey + '|' + str(ww.text)                                                                        
                            else:                                
                                emailkey = emailkey + '|' + str(w.text)

                        myjobdict[emailkey] = '| Job ' + y.tag[34:] + ' Differences'                        
                        emailkey = ''                
                #programdatalist
                elif y.tag[34:] == 'programdatalist': 
                    varcount = 0
                    for z in y:  #cycling through program data                                                    
                            varcount = varcount + 1
                            for w in z:                                 
                                if w.tag[34:] in ['notifyusers','notifydisplayusers']:
                                    for ww in w:                                        
                                        myjobdict['programdata' + str(varcount) + ' - ' + w.tag[34:] + '-' + ww.text] = ' | ProgramData Change'    
                                else:                                    
                                    myjobdict['programdata' + str(varcount) + ' - ' + w.tag[34:]] = w.text
                #customproperties
                elif y.tag[34:] in ['customproperties','envarlist','exitcodelist']: 
                    for z in y:                                                  
                        myjobdict[z[0].text + '-' + z[1].text] = ' | ' + y.tag[34:] + ' Change'                    
                #add successor dependancies
                elif y.tag[34:] in ['dependencies']:
                    if dependancycheck:
                        for z in y:                        
                            if z.tag[34:] in ['relcount']:
                                # add <relcount>
                                myjobdict[z.tag[34:]] = z.text                            
                            elif z.tag[34:] in ['relconditionlist']:
                                for w in z:
                                    dependancykey = ''
                                    for ww in w:                                    
                                        if dependancykey == '':
                                            dependancykey = dependancykey + ww.text + '-'
                                        else:
                                            dependancykey = dependancykey + ww.text                                
                                    myjobdict[dependancykey] = 'Job Dependancy Differences'                        
                else:
                    myjobdict[y.tag[34:]] = y.text       #add each tag with information to job dictionary
                                                            #note using [34:] to get past "{http://dto.wa.ca.com/application}" section in tag
            mydict[job_name] = myjobdict    #add job dictionary to full dictionary
        else:
            #capture app info to file if past section collecting
            if x.tag[34:] == 'job_templates':  #changed from script_lib                
                #mydict[appname] = myjobdict    06/14/22 - Changed to hard code 'APPLICATION" as key due to dev app naming is not same, corrects app compare settings
                mydict['APPLICATION'] = myjobdict
                needappdata = False                

            #Collect Application level information       
            if needappdata:
                if x.tag[34:] == 'defaults':                                 
                    for z in x: 
                        if z.tag[34:] == 'schedules':                    
                            for w in z:              
                                for ww in w:
                                    myjobdict[w.tag[34:] + '-' + ww.text] = '| App ' + w.tag[34:] + ' Differences'
                                
                        elif z.tag[34:] == 'rununit':
                            myjobdict[z[0].tag[34:]] = z[0].text
                        
                        #Emailing capture
                        elif z.tag[34:] in ['notifylist','alert_notifylist','snmp_notifylist']:
                            for w in z:
                                emailkey = ''                                                                        
                                for ww in w:
                                    if ww.tag[34:] == 'monitor_states' or ww.tag[34:] == 'mailaddresslist':
                                        for www in ww:
                                            emailkey = emailkey + '|' + str(www.text)                                                                        
                                    else:
                                        emailkey = emailkey + '|' + str(ww.text)
                                myjobdict[emailkey] = ' | App - ' + z.tag[34:] + ' Differences'
                                emailkey = ''
                
                #if tag is applopts collect app options
                elif x.tag[34:] == 'applopts':
                    for z in x:
                        myjobdict['App-applopts - ' + z.tag[34:]] = z.text
                
                #if tag is script_lib collect app scripts
                elif x.tag[34:] == 'script_lib':  #job_qualifier = x.attrib['name']
                    for z in x:
                        myjobdict['App-script_lib:' + z.attrib['name']] = z.text

                elif x.tag[34:] in ['version','comment']:            
                    myjobdict['App ' + x.tag[34:]] = x.text       #add each tag with app information to job dictionary

                elif x.tag[34:] in ['on_run','on_load']:
                    for z in x:
                        myjobdict['App - ' + x.tag[34:] + '| Javacall -' + x.tag[34:]] = z.text
    return mydict

def compareDictionaries (mydict, mydict2, prodfile, newfile, filename):
    #module will compare 2 dictionaries and print out the differences.
    #mydict = productions dictionary
    #mydict2 = new file dictionary
    #prodfile = path and file name for prod file location
    #newfile = path and file name for new file location
    #filename = just the prod file name
    job_defaults = {'notifynodefaults', 'snmp_notifynodefaults', 'alert_notifynodefaults', 'conditional', 'criticaljob', 'subappl_wait', 'hold', 'request','last_notify_email','last_notify_alerts','last_notify_snmp','estimate_endtime','prop_dueout','noinherit','notrigger_ifactive','suppress_nowork_notification','reason_required','App-applopts - estimate_endtime','App-applopts - prop_dueout','App-applopts - hold','App-applopts - noinherit','App-applopts - notrigger_ifactive','App-applopts - suppress_nowork_notification','App-applopts - reason_required'}
    nodifferences = True
    cngdict = {}    #dictionary to hold changed jobs
    newdict = {}    #dictionary to hold new jobs added
    deldict = {}    #dictionary to hold deleted jobs
    newaddjobs = {} #dictionary to hold new job information
    keycount = 100    #used for indexing key for same job

    print('Comparing Prod File:', prodfile, ' to New File:', newfile)
    #cycle through the dictionaries comparing Orig values to New
    for key in mydict:
        #check if job is in the production file            
        if key in mydict2:                    
            if mydict[key] == mydict2[key]:
                del mydict2[key]    #remove the key from New dicationary as they are the same            
            
            else: #cycle through all elements as there are differences
                nodifferences = False
                for elem in mydict[key]:
                    if elem in mydict2[key]:
                        if (mydict[key][elem] != mydict2[key][elem]) and ((elem in job_defaults) and (mydict[key][elem] == 'false')) != True:
                            keycount = keycount + 1                            
                            cngdict[filename + '-' + key + ' - Change ' + str(keycount)] = 'CHANGED Parameter: ' + elem + ': Orig: ' + mydict[key][elem] + ' New: ' + mydict2[key][elem]
                            print(key, elem + ': Orig:', mydict[key][elem], ' New:', mydict2[key][elem])
                            del mydict2[key][elem] #remove the element from New dicationary as we capture difference here            
                    else:
                        if (elem in job_defaults) and (mydict[key][elem] == 'false'):
                            #do not add or track the defaults as some xml files have them others do not
                            dave = 1
                        else:
                            keycount = keycount + 1
                            cngdict[filename + '-' + key + ' - Change ' + str(keycount)] = 'CHANGED Del Param: ' + elem + ': ' + mydict[key][elem]
                            print(key, elem + ':', mydict[key][elem], ' DELETED-not in new record')
        else: #job not found in prod file
            if 'App version' not in mydict[key]:
                nodifferences = False
                keycount = keycount + 1
                deldict[filename + '-' + key + ' - Change ' + str(keycount)] = 'DELETED Job'
                print(key,' job has been DELETED') #can remove from dictionary and save to another deleted dictionary            
    if nodifferences:
        print('No Differences Found in Files')

    print('')
    print('Comparing New File:', newfile, ' to Prod File:', prodfile)
    #cycle through the dictionaries comparing New values to Orig
    for key in mydict2:    
        if key in mydict:
            for elem in mydict2[key]:
                if elem not in mydict[key]:
                    if (elem in job_defaults) and (mydict2[key][elem] == 'false'):
                        #do not add or track the defaults as some xml files have them others do not
                        dave = 1
                    else:                        
                        nodifferences = False
                        keycount = keycount + 1
                        cngdict[filename + '-' + key + ' - Change ' + str(keycount)] = 'CHANGED Add Param: ' + elem  + ': ' + mydict2[key][elem]
                        print(key, elem + ': ', mydict2[key][elem], ' ADDED-not in orig record')
        else:
            if 'App version' not in mydict2[key]:
                nodifferences = False
                keycount = keycount + 1
                newdict[filename + '-' + key + ' - Change ' + str(keycount)] = 'ADDED Job'
                print(key,' job has been ADDED') #can move job information to a newjob dictionary for processing validation

                #add the new job information to new jobs dictionary for checking
                newaddjobs[filename + "-" + key] = mydict2[key]
                newaddjobs[filename + "-" + key]['JobName'] = key
                newaddjobs[filename + "-" + key]['AppName'] = filename

    if nodifferences:
        print('No Differences Found in Files')
    
    return cngdict, newdict, deldict, newaddjobs

def printappchanges(totalalldict):
    print('')
    print('Final Application Changes Listing')          
    sorted_dict = dict(sorted(totalalldict.items()))

    for key in sorted_dict:
        print(key, sorted_dict[key])

def clear():
    # define our clear function
    # import only system from os
    from os import system, name
    
    # for windows
    if name == 'nt':
        _ = system('cls')  
    # for mac and linux(here, os.name is 'posix')
    else:
        _ = system('clear')

def main():
    import xml.etree.ElementTree as ET
    import os
    finalcngdict = {}       #dictionary to hold changed jobs
    finalnewdict = {}       #dictionary to hold new jobs
    finaldeldict = {}       #dictionary to hold deleted jobs
    totalalldict = {}
    newappdict = {}         #dicationary to hold new apps added
    finalnewaddjobs = {}    #dictionary to hold new job information
    comparenewfile = ''     #used to hold a trimmed down new file name
    compareprodfile = ''    #used to hold a trimmed down prod  file name
    changeJiraNumber = ''   #used to hold the Change number or Jira story for file write
    continueProcessing = True
    processDependancys = True
    
    #clear the screen for fresh data
    clear()
    
    #get programming defaults
    xmlnewfilepath, xmlprodfilepath, processDependancys, printTooFile, saveFilePath = getDefaults()
    print('')
    changeJiraNumber = input('What is the Jira story or Change Request Number?: ')
    
    #change to Prod file directory and get file list
    os.chdir(xmlprodfilepath)
    prodfiles = os.listdir(xmlprodfilepath)
    #change to New file directory and get file list
    os.chdir(xmlnewfilepath)
    newfiles = os.listdir(xmlnewfilepath)

    print('')
    print('STARTING COMPAIRISONS')
    print('*****************************************************************************')
    for prodfile in prodfiles:
        continueProcessing = True
        if prodfile[-4:] == '.xml': #remove the end section
            compareprodfile = prodfile[0:-4]
        else:
            compareprodfile = prodfile  #6/28/22-Changed was compareprodfile

        for newfile in newfiles:            
            if (newfile[0:5].upper() == 'TEST_') or (newfile[0:5].upper() == 'VDEV_'):     #remove the prefix
                comparenewfile = newfile[5:]                
            elif newfile[0:4].upper() == 'E2E_':
                comparenewfile = newfile[4:]                
            else:
                comparenewfile = newfile
                
            if comparenewfile[-4:] == '.xml': #remove the end section
                comparenewfile = comparenewfile[0:-4]

            #print('prodfile=',compareprodfile)
            #print('newfile=',comparenewfile)
            #print('leng prod=',len(compareprodfile))
            #print('leng new=',len(comparenewfile))
            #print('length=',comparenewfile[len(compareprodfile):len(compareprodfile) + 1])
            if continueProcessing:
                #if compareprodfile == comparenewfile[0:len(compareprodfile)]:                                
                if ((compareprodfile == comparenewfile[0:len(compareprodfile)]) and (len(compareprodfile) == len(comparenewfile))) or ((compareprodfile == comparenewfile[0:len(compareprodfile)]) and (comparenewfile[len(compareprodfile):len(compareprodfile) + 1] == '_')):                
                    #fill prod dictionary            
                    mytree = ET.parse(os.path.join(xmlprodfilepath, prodfile))
                    root = mytree.getroot()
                    proddict = fillDictionaryFromXML(root, processDependancys)
                    
                    #fill new dictionary            
                    mytree = ET.parse(os.path.join(xmlnewfilepath, newfile))
                    root = mytree.getroot()    
                    newdict = fillDictionaryFromXML(root, processDependancys)                          
                    #print('proddict=',proddict)
                    #print('newdict=',newdict)
                    #compare the dictionaries  
                    cngdict, newdict, deldict, newaddjobs = compareDictionaries(proddict, newdict, xmlprodfilepath + '\\' + prodfile, xmlnewfilepath + '\\' + newfile, prodfile)
                    
                    #accumulate all changes into 1 file
                    finalcngdict.update(cngdict)
                    finalnewdict.update(newdict)                    
                    finaldeldict.update(deldict)
                    finalnewaddjobs.update(newaddjobs)

                    #remove files processed
                    newfiles.remove(newfile)                    
                    continueProcessing = False

        if continueProcessing:        
            print(newfile, 'File not found in PROD folder')
        print('*****************************************************************************')

    for file in newfiles:
        print('NEW Application, not found in Prod file folder:', file)        
        #Accumulate the jobs to newjobs dictionary                    
        mytree = ET.parse(os.path.join(xmlnewfilepath, file))
        root = mytree.getroot()    
        newdict = fillDictionaryFromXML(root, processDependancys)
                
        newaddjobs = {}
        for job in newdict:
            #add the new job information to new jobs dictionary for checking                        
            newaddjobs[file + "-" + job] = newdict[job]                      
            newaddjobs[file + "-" + job]['AppName'] = file
            newaddjobs[file + "-" + job]['JobName'] = job
            if "Job_Type" not in newdict[job]:
                newaddjobs[file + "-" + job]['Job_Type'] = 'New Application'

        #add description to changes information for the new application
        newappdict[file] = 'NEW Application'
        #newappdict['Job_Type'] = 'NEW Application'
        finalnewdict.update(newappdict)

        finalnewaddjobs.update(newaddjobs)
    
    print('*****************************************************************************')
    
    #accumlate all changes into 1 dictionary
    totalalldict.update(finalcngdict)
    totalalldict.update(finalnewdict)
    totalalldict.update(finaldeldict)
    #print final result
    printappchanges(totalalldict)

    #run job checking routine    
    finalnewaddjobs = checknewjobsdata(finalnewaddjobs)

    #print report
    if printTooFile == 'Excel':
        #print to excel final changes for differences
        printexcelfile(totalalldict, 1, changeJiraNumber, saveFilePath)
        printexcelfile(finalnewaddjobs, 2, changeJiraNumber, saveFilePath)
    else:
        #print to text file
        printtextfile(totalalldict, 1, changeJiraNumber, saveFilePath)    
        printtextfile(finalnewaddjobs, 2, changeJiraNumber, saveFilePath)
    
    #print('finaladdjobs=', finalnewaddjobs)
    #print('finalcngdict=', finalcngdict)
    #print('finaldeldict=', finaldeldict)
    os.close

def printexcelfile(totalalldict, changetype, changeJiraNumber, saveFilePath):
    #totalalldict = dictionary of changes coming in to print
    #changetype(1=All Differences,2=New Jobs)
    #changeJiraNumber = Used in construction of file saving    
    #saveFilePath = path to where results file will be saved

    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter  #utility to get column letters.
    from openpyxl.styles import Font, PatternFill    
    import re

    #used to color row for task/link jobs to green
    colornextrowgreen = False
    colornextrowblue = False      
    colornextrowpurple = False       
    linktaskjob = False
    externaljob = False
    applicationrow = False
    fillcolorgreen = PatternFill(patternType='solid', fgColor='D4FFD1')
    fillcolordkgreen = PatternFill(patternType='solid', fgColor='006400')
    fillcolorred = PatternFill(patternType='solid', fgColor='FFDDD1')
    fillcolorltblue = PatternFill(patternType='solid', fgColor='D1EFFF')
    fillcolorltpurple = PatternFill(patternType='solid', fgColor='E6BBF6')
    fillcoloryellow = PatternFill(patternType='solid', fgColor='FFF700')
    
    workbookfile = saveFilePath + '\\Comparison_Results-' + changeJiraNumber + '.xlsx'

    if changetype == 1:
        #create new empty work book
        wb = Workbook()
    else:
        #load existing work to add new jobs information
        wb = load_workbook(workbookfile)

    #set to first sheet to begin working on
    ws = wb.active

    ws.title = "App Differences"    #name the sheet
    headings = []    
    if changetype == 1:        
        headings.append('Final Application Changes Listing')
        ws.append(headings)              
        sorted_dict = dict(sorted(totalalldict.items()))

        for key in sorted_dict:
            headings = []
            headings = [key] + [sorted_dict[key]]            
            ws.append(headings)
    else:
        ws.append([])        
        ws.append([])
        ws.append(['NEW JOBS TO VALIDATE'])                
        ws.append(['    APP INFORMATION'])                
        ws.append(['     LINK/TASK JOBS'])                
        ws.append(['     EXTER MON JOBS'])                
        sorted_dict = dict(sorted(totalalldict.items()))
        for key in sorted_dict: #cycle through all rows in the added jobs dictionary
            #if 'INVALID_ENTRIES' in totalalldict[key]:
            headings = ['AppName', 'JobName', 'INVALID_ENTRIES'] + list(totalalldict[key].keys())                                          #create column headings using the data. all keys same            
            ws.append(headings)                                                                                         #add the heading to sheet            
            grades = [totalalldict[key]['AppName'], totalalldict[key]['JobName'], totalalldict[key]['INVALID_ENTRIES']] + list(totalalldict[key].values())    #get all the grades as a list for each person
            ws.append(grades)                                                                                           #append to the excel sheet each row of data
            ws.append([])                                                                                               #add a blank line
            #else:
            #    headings = ['AppName', 'JobName'] + list(totalalldict[key].keys())                                          #create column headings using the data. all keys same            
            #    ws.append(headings)                                                                                         #add the heading to sheet            
            #    grades = [totalalldict[key]['AppName'], totalalldict[key]['JobName']] + list(totalalldict[key].values())    #get all the grades as a list for each person
            #    ws.append(grades)                                                                                           #append to the excel sheet each row of data
            #    ws.append([])                                                                                               #add a blank line                

    #save workbook
    wb.save(workbookfile)

    #save workbook
    #trySaveFile = True
    #while trySaveFile:
    #    try:
    #        wb.save(workbookfile)            
    #        print('Successfully saved')
    #        trySaveFile = False
    #    except IOError as err:
    #        hold = input('Write permissions to file, close file and try again.  Enter any key after closing file.')
    #        #continue       
    

    #always close the workbook
    wb.close

    #lets add some color and bolding
    wb = load_workbook(workbookfile)
    ws = wb.active
    for row in ws.iter_rows():         
        makecellbold = False      
        
        if colornextrowgreen:
            linktaskjob = True
        else:
            linktaskjob = False

        if colornextrowblue:
            externaljob = True
        else:
            externaljob = False
        
        if colornextrowpurple:
            applicationrow = True
        else:
            applicationrow = False

        colornextrowgreen = False
        colornextrowblue = False       
        colornextrowpurple = False

        for cell in row:

            if cell.column == 2 and ws[cell.coordinate].value in ['NEW Application', 'ADDED Job']:
                ws[str(cell.coordinate)].font = Font(bold='true', color='006400')
            if cell.column == 2 and ws[cell.coordinate].value in ['CHANGED Add Param: donotrun: true']:
                ws[str(cell.coordinate)].fill = fillcolorred
                ws[str(cell.coordinate)].font = Font(bold='true', color='00FF0000')
            elif cell.column == 2 and ws[cell.coordinate].value == 'DELETED Job':
                ws[str(cell.coordinate)].font = Font(bold='true', color='0B22F6')
                
            #if its column 1 and has 1 of those values then make every cell in this row bold        
            if cell.column == 1 and (ws[cell.coordinate].value == 'AppName' or ws[cell.coordinate].value == 'JobName'):            
                makecellbold = True
            if ws[cell.coordinate].value == 'Final Application Changes Listing':
                ws[str(cell.coordinate)].font = Font(bold='true')
            if ws[cell.coordinate].value == 'NEW JOBS TO VALIDATE':
                ws[str(cell.coordinate)].fill = fillcoloryellow
                ws[str(cell.coordinate)].font = Font(bold='true')
            if ws[cell.coordinate].value == '    APP INFORMATION':
                ws[str(cell.coordinate)].fill = fillcolorltpurple
                ws[str(cell.coordinate)].font = Font(bold='true')
            if ws[cell.coordinate].value == '     LINK/TASK JOBS':                
                ws[str(cell.coordinate)].fill = fillcolorgreen
                ws[str(cell.coordinate)].font = Font(bold='true')
            if ws[cell.coordinate].value == '     EXTER MON JOBS':
                ws[str(cell.coordinate)].fill = fillcolorltblue  
                ws[str(cell.coordinate)].font = Font(bold='true')              
            #if jobvalidation and cell.column == 3 and ws[cell.coordinate].value not in ['NONE', 'LINK/TASK JOBS']:
            if cell.column == 3 and ws[cell.coordinate].value not in ['NONE']:
                ws[str(cell.coordinate)].font = Font(bold='true', color='00FF0000')
            if (cell.column == 3) and (ws[cell.coordinate].value == 'Job_Type') or ((cell.column == 4) and (ws[cell.coordinate].value == 'Job_Type')):
                checkposforjobNUM = re.findall("\d+", cell.coordinate)[0]                
                if (ws['D' + str(int(checkposforjobNUM) + 1)].value in ['link','task']): # or (ws['C' + str(int(checkposforjobNUM) + 1)].value in ['link','task']):
                    colornextrowgreen = True
                if (ws['D' + str(int(checkposforjobNUM) + 1)].value in ['ext_job','extmon_job']): 
                    colornextrowblue = True
            if (cell.column == 4) and (ws[cell.coordinate].value == 'App version'):
                colornextrowpurple = True
            #continue making all other cells in the header row bold
            if makecellbold:
                ws[str(cell.coordinate)].font = Font(bold='true')
            if linktaskjob:                
                ws[str(cell.coordinate)].fill = fillcolorgreen
            if externaljob:
                ws[str(cell.coordinate)].fill = fillcolorltblue
            if applicationrow:
                ws[str(cell.coordinate)].fill = fillcolorltpurple

    #save final workbook
    wb.save(workbookfile)

    #save final workbook
    #trySaveFile = True
    #while trySaveFile:
    #    try:
    #        wb.save(workbookfile)            
    #        print('Successfully saved')
    #        trySaveFile = False
    #    except IOError as err:
    #        hold = input('Write permissions to file, close file and try again.  Enter any key after closing file.')
    #        #continue  

    #always close the workbook
    wb.close

def printtextfile(totalalldict, changetype, changeJiraNumber, saveFilePath):
    #totalalldict = dictionary of changes coming in to print
    #changetype(1=All Differences,2=New Jobs)
    #changeJiraNumber = Used in construction of file saving    
    #saveFilePath = path to where results file will be saved
    
    import csv
    workbookfile = saveFilePath + '\\Comparison_Results-' + changeJiraNumber + '.txt'
    #workbookfile = 'C:\\CAWACompare\\Results\\Comparison_Results-' + changeJiraNumber + '.txt'    

    sorted_dict = dict(sorted(totalalldict.items()))    
    if changetype == 1:
        with open(workbookfile, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            xml_data = []
            xml_data.append('Final Application Changes Listing')
            writer.writerow(xml_data)
            for key in sorted_dict:
                xml_data = []
                xml_data.append(key)
                xml_data.append(sorted_dict[key])
                writer.writerow(xml_data)
            writer.writerow('')
            writer.writerow('')
            xml_data = []
            xml_data.append('NEW JOBS TO VALIDATE')
            writer.writerow(xml_data)
    else:        
        with open(workbookfile, 'a', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            for key in sorted_dict:
                xml_key_data = []
                xml_key_value = []
                firstelement = True            
                for element in sorted_dict[key]:
                    if firstelement:
                        xml_key_data.append('AppName')
                        xml_key_value.append(sorted_dict[key]['AppName']) 
                        xml_key_data.append('JobName')
                        xml_key_value.append(sorted_dict[key]['JobName'])
                        xml_key_data.append('INVALID_ENTRIES')
                        xml_key_value.append(sorted_dict[key]['INVALID_ENTRIES'])
                        firstelement = False

                    xml_key_data.append(element)
                    xml_key_value.append(sorted_dict[key][element])

                writer.writerow(xml_key_data)
                writer.writerow(xml_key_value)
                writer.writerow('')

    #always close the file
    csvfile.close()

def getfilepath(filePath, fileSource):    
    newFilePath = filePath
    if fileSource == 'New':
        fileSourcePath = 'New File Path'        
    elif fileSource == 'Prod':
        fileSourcePath = 'Prod File Path'
    elif fileSource == 'Save':
        fileSourcePath = 'Save Results File Path'
    
    filePathAnswer = input('Current ' + fileSourcePath + ' is set too: ' + filePath + ' Keep this path? (Y) Change this path? (C):')
    while (filePathAnswer.upper() != 'Y'):
        filePathAnswer = input('What is your ' + fileSourcePath + ' location?:')
        newFilePath = filePathAnswer
        filePathAnswer = input('Current ' + fileSourcePath + ' is set too: ' + newFilePath + ' Keep this path? (Y) Change this path? (C):')
        
    #return the file path
    return newFilePath

def getDefaults():
    # importing the module
    import json
    changeOption = "displaydefaults"

    # get the default information
    with open('C:\\CAWACompare\\bin\\CAWACompareSettings.json') as json_file:
        data = json.load(json_file)
        newFilePath = data['newFilesFolder']
        prodFilePath = data['prodFilesFolder']
        dependancycheck = data['CheckDependanys']
        printTooFile = data['FilePrintType']
        saveFilePath = data['FileSavePath']

    while changeOption != "":        
        #clear the screen for fresh data
        clear()
        #Display all Options Using
        print('Default Options Using')
        print('1. New      Files Folder: ', newFilePath)
        print('2. Prod     Files Folder: ', prodFilePath)
        print('3. Results  Files Folder: ', saveFilePath)
        print('4. Check For Dependancys: ', dependancycheck)
        print('5. Save as (Excel/Text) : ', printTooFile)
        print('')
        changeOption = input('These are your default settings.  If you like to change any settings enter selection Number or Enter to Use defaults:')

        if changeOption == '1':
            # Check for new file path
            newFilePath = getfilepath(newFilePath, 'New')
        if changeOption == '2':
            # Check for prod file path
            prodFilePath = getfilepath(prodFilePath, 'Prod')
        if changeOption == '3':
            saveFilePath = getfilepath(saveFilePath, 'Save')
        if changeOption == '4':
            # Check for job naming standards
            processDependancys = input('Check Jobs Dependancys is set too (' + str(data['CheckDependanys']) + ')  Keep this setting? (Y) Change this setting? (C):')
            if processDependancys in ['Y', 'y']:
                dependancycheck = data['CheckDependanys']
            else:
                dependancycheck = not data['CheckDependanys']
        if changeOption == '5':
            # Check for file print type
            if printTooFile == 'Excel':
                processDependancys = input('Results printed to (' + data['FilePrintType'] + ') File.  Keep this setting? (Y) or Print to Text File? (T):')
                if processDependancys in ['Y', 'y']:
                    printTooFile = data['FilePrintType']
                else:
                    printTooFile = 'Text'
            else:
                processDependancys = input('Results printed to (' + data['FilePrintType'] + ') File.  Keep this setting? (Y) or Print to Excel File? (E):')
                if processDependancys in ['Y', 'y']:
                    printTooFile = data['FilePrintType']
                else:
                    printTooFile = 'Excel'
    
    # if paths changed update back to json file
    if (newFilePath != data['newFilesFolder']) or (prodFilePath != data['prodFilesFolder']) or (saveFilePath != data['FileSavePath']) or (dependancycheck != data['CheckDependanys']) or (printTooFile != data['FilePrintType']):
        data['newFilesFolder'] = newFilePath
        data['prodFilesFolder'] = prodFilePath
        data['FileSavePath'] = saveFilePath
        data['CheckDependanys'] = dependancycheck
        data['FilePrintType'] = printTooFile
        
        with open("C:\\CAWACompare\\bin\\CAWACompareSettings.json", "w") as outfile:
            json.dump(data, outfile)
        
    return newFilePath, prodFilePath, dependancycheck, printTooFile, saveFilePath

def checknewjobsdata(testdic):
    jobname = ''
        
    #get sub-system listing
    List = open("C:\\CAWACompare\\bin\\SubSystemsList.txt").read().splitlines()

    for key in testdic:            
        validJobName = True
        validOmniJobName = False
        errormsg = ''        
        if ('App version' in testdic[key]) or (testdic[key]['Job_Type'].upper() == 'LINK') or (testdic[key]['Job_Type'].upper() == 'TASK'):                            
            testdic[key]['INVALID_ENTRIES'] = 'NONE'
        elif (testdic[key]['Job_Type'].upper() == 'EXT_JOB') or (testdic[key]['Job_Type'].upper() == 'EXTMON_JOB'):
            #place holder, add check for the scheduled limits
            #<app:scheduled>00:00 TODAY</app:scheduled> From value
            #<app:scheduledlimit>00:00 TOMORROW</app:scheduledlimit> Too Value
            if ('scheduled' not in testdic[key]) or ('scheduledlimit' not in testdic[key]):
                errormsg = errormsg + ' Invalid Scheduled Limits'
        else:
            #get the job name without qualifier
            if "." in testdic[key]['JobName']:
                jobname = testdic[key]['JobName'][:testdic[key]['JobName'].find('.')]      

            # Omni name job validation
            if jobname[0:4].upper() in ['PAM_','PPC_']:
                validOmniJobName = True
            elif jobname[0:5].upper() in ['PAXE_','PICM_','PMDS_','PEMR_','PESP_','PINV_','PMCM_','POMS_','PPAL_','PPUB_','PSDC_','PWMS_']:
                validOmniJobName = True
            elif jobname[0:6].upper() in ['BOPIS_','PCUST_','PITEM_','PXREF_']:
                validOmniJobName = True
            elif jobname[0:7].upper() in ['PACMAN_','PCLEAR_','PHBASE_','PNGPOS_','PPURGE_','PTOOLS_','PTRANS_']:
                validOmniJobName = True
            elif jobname[0:8].upper() in ['PREPORT_','PSELECT_','PROCESS_']:
                validOmniJobName = True
            elif jobname[0:9].upper() in ['PAUTOPOP_','PCATALOG_','PHYSICAL_','PPRICECA_','PPRICEJP_','PPRICEUK_','PPRICING_','PPRODUCT_','PPUBLISH_']:
                validOmniJobName = True
            elif jobname[0:11].upper() in ['POSRECLAIM_']:
                validOmniJobName = True
            
            if not validOmniJobName:
                # check for prod
                if jobname[0:1].upper() != 'P':
                    errormsg = 'First letter of Job name not P'

                # check for subsystem name
                if jobname[1:4] not in List:
                    errormsg = errormsg + ' Invalid SubSystem Name'

                # check for frequency: Frequency (D-Daily, H-Hourly, W-Weekly, M-monthly, R-Request etc)
                if jobname[len(jobname)-2] not in ['D','H','W','M','R','U']: 
                    errormsg = errormsg + ' Invalid Frequency'

                #check for Last Character: Platform (U-Unix, L-Linux, O-Oracle, R–Request, P-Peoplesoft, N/W-Windows, S-MSSQL Job, ? question about these UP-SFTP Upload, DW-SFTP Download etc.)
                if (jobname[-2:].upper() in ['UP', 'DW']) and (testdic[key]['Job_Type'].upper() != 'SFTP_JOB'):
                    validJobName = False
                elif (jobname[-1].upper() in ['U', 'L', 'R']) and (jobname[-2].upper() not in ['UP', 'DW']) and (testdic[key]['Job_Type'].upper() != 'UNIX_JOB'):      
                    validJobName = False
                elif (jobname[-1].upper() == 'O') and ((testdic[key]['Job_Type'].upper() != 'ORACLE_APPS_REQUEST_SET') and (testdic[key]['Job_Type'].upper() != 'ORACLE_APPS_SINGLE_REQUEST')):
                    validJobName = False
                elif (jobname[-1].upper() == 'P') and (jobname[-2:].upper() not in ['UP', 'DW']) and (testdic[key]['Job_Type'].upper() != 'PEOPLESOFT_JOB'):
                    validJobName = False
                elif ((jobname[-1].upper() in ['N', 'W']) and (testdic[key]['Job_Type'].upper() != 'NT_JOB')) and (jobname[-2:].upper() not in ['UP', 'DW']):
                    validJobName = False
                elif (jobname[-1].upper() == 'S') and (testdic[key]['Job_Type'].upper() != 'MSSQLSERVER_JOB'):                
                    validJobName = False
                elif (jobname[-1].upper() not in ['U', 'L', 'R','O','P','N','W','S']):
                    validJobName = False
                if not validJobName:    
                    errormsg = errormsg + ' Invalid Platform'
            
        if 'donotrun' in testdic[key]:
            errormsg = errormsg + ' Job has donotrun setting'
            #print('value=',testdic[key]['donotrun'])

        if errormsg != '':
            testdic[key]['INVALID_ENTRIES'] = errormsg
        elif errormsg == '':
            testdic[key]['INVALID_ENTRIES'] = 'NONE'
        elif not validOmniJobName:
            testdic[key]['INVALID_ENTRIES'] = 'Invalid Omni Job Name'            

    return testdic

main()

